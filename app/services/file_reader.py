"""File reader service — extract text from common office file types."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path


@dataclass
class FileContent:
    text: str
    filename: str
    file_type: str       # extension including dot: ".md", ".pdf"
    file_hash: str       # SHA-256 hex digest
    file_size: int
    modified_at: datetime  # file mtime as tz-aware UTC datetime


_MAX_TEXT_LENGTH = 8000


def _truncate(text: str) -> str:
    """Truncate text to _MAX_TEXT_LENGTH chars, appending a note if truncated."""
    total_chars = len(text)
    if total_chars <= _MAX_TEXT_LENGTH:
        return text
    return text[:_MAX_TEXT_LENGTH] + f"... [truncated, {total_chars} chars total]"


class FileReaderService:
    SUPPORTED_TYPES = {".md", ".txt", ".docx", ".pdf", ".xlsx"}

    @staticmethod
    def read_file(path: Path) -> FileContent:
        """Dispatch to format-specific reader. Raises ValueError for unsupported types."""
        path = Path(path)
        ext = path.suffix.lower()

        if ext not in FileReaderService.SUPPORTED_TYPES:
            raise ValueError(
                f"Unsupported file type: '{ext}'. "
                f"Supported types: {sorted(FileReaderService.SUPPORTED_TYPES)}"
            )

        file_hash = FileReaderService._compute_hash(path)
        stat = path.stat()
        file_size = stat.st_size
        modified_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)

        readers = {
            ".md": FileReaderService._read_markdown,
            ".txt": FileReaderService._read_plaintext,
            ".docx": FileReaderService._read_docx,
            ".pdf": FileReaderService._read_pdf,
            ".xlsx": FileReaderService._read_xlsx,
        }

        text = readers[ext](path)
        text = _truncate(text)

        return FileContent(
            text=text,
            filename=path.name,
            file_type=ext,
            file_hash=file_hash,
            file_size=file_size,
            modified_at=modified_at,
        )

    @staticmethod
    def _compute_hash(path: Path) -> str:
        """SHA-256 of file content."""
        sha256 = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                sha256.update(chunk)
        return sha256.hexdigest()

    @staticmethod
    def _read_markdown(path: Path) -> str:
        """Read .md as UTF-8 text."""
        return path.read_text(encoding="utf-8")

    @staticmethod
    def _read_plaintext(path: Path) -> str:
        """Read .txt with encoding fallback: UTF-8 -> GBK -> latin-1."""
        for encoding in ("utf-8", "gbk", "latin-1"):
            try:
                return path.read_text(encoding=encoding)
            except (UnicodeDecodeError, ValueError):
                continue
        # latin-1 should never fail, but just in case
        return path.read_text(encoding="latin-1", errors="replace")

    @staticmethod
    def _read_docx(path: Path) -> str:
        """Extract text from .docx using python-docx.
        Import inside function to handle missing dependency gracefully.
        Return error message if python-docx not installed.
        """
        try:
            from docx import Document  # type: ignore[import-untyped]
        except ImportError:
            return "[Error: python-docx not installed, run: pip install python-docx]"

        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)

    @staticmethod
    def _read_pdf(path: Path) -> str:
        """Extract text from .pdf using PyPDF2.
        Import inside function to handle missing dependency gracefully.
        Return error message if PyPDF2 not installed.
        """
        try:
            from PyPDF2 import PdfReader  # type: ignore[import-untyped]
        except ImportError:
            return "[Error: PyPDF2 not installed, run: pip install PyPDF2]"

        reader = PdfReader(str(path))
        pages_text = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages_text.append(text)
        return "\n".join(pages_text)

    @staticmethod
    def _read_xlsx(path: Path) -> str:
        """Extract text from .xlsx using openpyxl.
        Read sheet by sheet, row by row, format as tab-separated text.
        Import inside function to handle missing dependency gracefully.
        Return error message if openpyxl not installed.
        """
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError:
            return "[Error: openpyxl not installed, run: pip install openpyxl]"

        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines)
